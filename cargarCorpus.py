from openpyxl import load_workbook
import datetime
import pandas as pd

class Tweet:
	def __init__(self, nombre, partido, txtOriginal, txtToken, txtStem, urls, emojis, menciones, hashtags, esRT, nRT, nFav, fecha):
		
		self.nombre		= nombre
		self.partido	= partido
		self.txtOriginal= txtOriginal
		self.txtToken	= txtToken
		self.txtStem	= txtStem
		self.urls		= urls
		self.emojis 	= emojis
		self.menciones 	= menciones
		self.hashtags 	= hashtags
		self.esRT 		= esRT
		self.nRT 		= nRT
		self.nFav 		= nFav
		self.fecha 		= fecha

def cargarPandas(nCuentas = 121):

	corpus = {}
	excels = []

	lista = load_workbook(filename='Cuentas.xlsx')['Hoja1']

	for politico in lista.iter_rows(min_row = 1, max_col = 3, max_row = nCuentas):
		if politico[1].value is not None:
			ex = (pd.read_excel('./tweetsProcesados/' + politico[2].value + '/' + politico[0].value + '/tuits - ' + politico[0].value + '.xlsx', 
				encode = 'utf-8', sheet_name = [0,1,2,3,4,5], error_bad_lines = False, header = None,
				names = ['tweet_completo', 'esRT', 'nRT', 'nFavs', 'dia', 'hora', 'minuto', 'tokens', 'lexemas', 'urls', 'emojis', 'menciones', 'hashtags']))
			for i in range(0,6):
				try:
					ex[i]['partido'] = politico[2].value
				except:
					pass
				try:	
					ex[i]['nombre'] = politico[0].value
				except:
					pass	
				try:	
					ex[i].esRT = [True if c == 'Y' else False for c in ex[i].esRT]
				except:
					pass	
				try:	
					ex[i].nRT  = [int(n) for n in ex[i].nRT]
				except:
					pass	
				try:	
					ex[i].nFavs  = [int(n) for n in ex[i].nFavs]
				except:
					pass	
				try:	
					ex[i].dia  = [int(n) for n in ex[i].dia]
				except:
					pass	
				try:	
					ex[i].hora  = [int(n) for n in ex[i].hora]
				except:
					pass	
				try:	
					ex[i].minuto  = [int(n) for n in ex[i].minuto]
				except:
					pass	
				try:	
					ex[i].hashtags  = [e[1].split() if type(e[1]) == str else [] for e in ex[i].hashtags.iteritems()]
				except:
					pass	
				try:	
					ex[i].tokens  = [e[1].split() if type(e[1]) == str else [] for e in ex[i].tokens.iteritems()]
				except:
					pass	
				

			excels.append(ex)

	corpus['enero'] 	= pd.concat([e[0] for e in excels])
	corpus['febrero'] 	= pd.concat([e[1] for e in excels])
	corpus['marzo'] 	= pd.concat([e[2] for e in excels])
	corpus['abril'] 	= pd.concat([e[3] for e in excels])
	corpus['mayo'] 		= pd.concat([e[4] for e in excels])
	corpus['junio']		= pd.concat([e[5] for e in excels])

	return corpus


def cargarPorFecha(nCuentas = 121, fechaMin = datetime.datetime(2020,1,1), fechaMax = datetime.datetime.now()):
	corpus = {}
	lista = load_workbook(filename='Cuentas.xlsx')['Hoja1']

	for politico in lista.iter_rows(min_row = 1, max_col = 3, max_row = nCuentas):
		
		if politico[1].value is not None:
			wb = load_workbook(filename='./tweetsProcesados/' + politico[2].value + '/' + politico[0].value + '/tuits - ' + politico[0].value + '.xlsx')
			for ws in wb:
				for tweet in ws.iter_rows():
					fecha = datetime.datetime(2020, int(ws.title), int(tweet[4].value))	#Año - mes - dia
					if fecha >= fechaMin and fecha <= fechaMax:
						for i in range(7,13):
								if tweet[i].value is None:
									tweet[i].value = ''

						if fecha in corpus: #Ya habia al menos 1 tw con esa fecha

							t = Tweet(politico[0].value, politico[2].value,	tweet[0].value, 
								(tweet[7].value).split(), (tweet[8].value).split(), (tweet[9].value).split(),
								(tweet[10].value).split(), (tweet[11].value).split(), (tweet[12].value).split(), 
								tweet[1].value == 'Y', int(tweet[2].value), int(tweet[3].value), fecha)
							
							corpus[fecha].append(t)

						else:
							corpus[fecha] = []

							t = Tweet(politico[0].value, politico[2].value,	tweet[0].value, 
								(tweet[7].value).split(), (tweet[8].value).split(), (tweet[9].value).split(),
								(tweet[10].value).split(), (tweet[11].value).split(), (tweet[12].value).split(), 
								tweet[1].value == 'Y', int(tweet[2].value), int(tweet[3].value), fecha)
							
							corpus[fecha].append(t)

	return corpus

def cargarPorFechaFiltrarPartido(partidos, nCuentas = 121, fechaMin = datetime.datetime(2020,1,1), fechaMax = datetime.datetime.now()):
	corpus = {}
	lista = load_workbook(filename='Cuentas.xlsx')['Hoja1']

	for politico in lista.iter_rows(min_row = 1, max_col = 3, max_row = nCuentas):
		
		if politico[2].value not in partidos:
			continue

		if politico[1].value is not None:
			wb = load_workbook(filename='./tweetsProcesados/' + politico[2].value + '/' + politico[0].value + '/tuits - ' + politico[0].value + '.xlsx')
			for ws in wb:
				for tweet in ws.iter_rows():
					fecha = datetime.datetime(2020, int(ws.title), int(tweet[4].value))	#Año - mes - dia
					if fecha >= fechaMin and fecha <= fechaMax:
						for i in range(7,13):
								if tweet[i].value is None:
									tweet[i].value = ''

						if fecha in corpus: #Ya habia al menos 1 tw con esa fecha

							t = Tweet(politico[0].value, politico[2].value,	tweet[0].value, 
								(tweet[7].value).split(), (tweet[8].value).split(), (tweet[9].value).split(),
								(tweet[10].value).split(), (tweet[11].value).split(), (tweet[12].value).split(), 
								tweet[1].value == 'Y', int(tweet[2].value), int(tweet[3].value), fecha)
							
							corpus[fecha].append(t)

						else:
							corpus[fecha] = []

							t = Tweet(politico[0].value, politico[2].value,	tweet[0].value, 
								(tweet[7].value).split(), (tweet[8].value).split(), (tweet[9].value).split(),
								(tweet[10].value).split(), (tweet[11].value).split(), (tweet[12].value).split(), 
								tweet[1].value == 'Y', int(tweet[2].value), int(tweet[3].value), fecha)
							
							corpus[fecha].append(t)

	return corpus

def filtrarCorpusPartidos(corpus, partidos):
	for dia in corpus:
		corpus[dia] = [t for t in corpus[dia] if t.partido in partidos]
	return corpus

