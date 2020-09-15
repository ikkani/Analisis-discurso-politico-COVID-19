import os
import tweepy
from openpyxl import load_workbook, Workbook
import datetime

def descargar():
	auth = tweepy.OAuthHandler("U9rgq3K0kWPUm3rYurEZgku1o", "bGAWJKD8gJ55EpufWVQz22ALOQRpoXsEXqxPpdsfvhQMeLtesU")
	access_token = "331544241-fYorH0zGtpdX0pOEEYD5zO9tq7dX2HheLnJLeghg"
	access_token_secret = "SnzuLkNfs7FL3mEHBlnRWYc8EuxE2GWoHdROOcZbbHwwT"

	auth.set_access_token(access_token, access_token_secret)
	api = tweepy.API(auth)

	wbCuentas = load_workbook(filename='Cuentas.xlsx')
	sheet = wbCuentas['Hoja1']

	parties = ["Cs", "ERC", "JpC", "PNV", "PP", "PRC", "PSOE", "UP", "VOX"]
	nCuentas = 121

	for party in parties:
		if not os.path.exists('./root/' + party):
			os.mkdir('./root/' + party)


	for politico in sheet.iter_rows(min_row = 1, max_col = 3, max_row = nCuentas):
		if politico[1].value is not None: #Tiene tw
			nTweets = 0
			if not os.path.exists('./root/' + politico[2].value + '/' + politico[0].value): #Creacion carpeta
				os.mkdir('./root/' + politico[2].value + '/' + politico[0].value)

			if not os.path.exists('./root/' + politico[2].value + '/' + politico[0].value + '/tuits - ' + politico[0].value + '.xlsx'): #WB
				print(politico[0].value + " no tenia xlsx, creando...")
				wb = Workbook()
				del wb["Sheet"]
				wb.create_sheet("1")	#Enero
				wb.create_sheet("2")	#Febrero
				wb.create_sheet("3")	#Marzo
				wb.create_sheet("4")	#Abril
				wb.create_sheet("5")	#Mayo
				wb.save('./root/' + politico[2].value + '/' + politico[0].value + '/tuits - ' + politico[0].value + '.xlsx')
			else:
				wb = load_workbook(filename='./root/' + politico[2].value + '/' + politico[0].value + '/tuits - ' + politico[0].value + '.xlsx')
				wb.create_sheet("6")	#Junio

			for tweet in tweepy.Cursor(api.user_timeline, id = politico[1].value, tweet_mode= 'extended').items():
				if (tweet.created_at.year < 2020): #Comprobamos que sea de este año
					break
				if (tweet.lang != 'es'): #Comprobamos que sea en español
					continue

				ws = wb[str(tweet.created_at.month)]

				rep = False
				twId = 'id' + tweet.id_str
				for idExcel in ws["J"]:
					if (idExcel.value == twId):
						rep = True
						break
				if (rep):
					break 
				
				if ('retweeted_status' in tweet._json): #Es un rt
					text = tweet.retweeted_status.full_text
					rtText = 'Y'
				else:
					text = tweet.full_text
					rtText = 'N'

				if tweet.place is None:
					city = ''
					coordinates = ''
				else:
					city = tweet.place.name
					coordinates = str(tweet.place.bounding_box.coordinates)

				ws.append([text, rtText, tweet.retweet_count, tweet.favorite_count, tweet.created_at.day, tweet.created_at.hour, tweet.created_at.minute, city, coordinates, twId])
				nTweets = nTweets + 1
			print ("[X] " + politico[0].value + ". Número de tuits nuevos: " + str(nTweets))
			wb.save('./root/' + politico[2].value + '/' + politico[0].value + '/tuits - ' + politico[0].value + '.xlsx')


