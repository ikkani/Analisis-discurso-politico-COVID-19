from openpyxl import load_workbook, Workbook
import datetime
import re
from nltk.corpus import stopwords
from nltk.stem import SnowballStemmer
import emoji
import os
from spacy.lang.es.stop_words import STOP_WORDS

###################PREPROCESSING###################
stop_words = set(stopwords.words('spanish'))
stemmer = SnowballStemmer('spanish')

def process_urls(string):
	pat_url = re.compile("https{0,1}://t\\.co/\w{6,10}")
	url = pat_url.findall(string)
	return [pat_url.sub(' ', string), url]

def process_emojis(string):
    return [emoji.get_emoji_regexp().sub(' ', string), emoji.get_emoji_regexp().findall(string)]

def remove_newline(string):
	return string.replace("\n", " ")

def remove_punctuation(string):
	symbols = ["." ,"," ,"¿" ,"?" ,"¡" ,"!" ,";" ,":" ,"'", " - " ,"\\" ,"(" ,")" ,"<" ,">" ,"=" ,"[" ,"]" ,"^" ,"`" ,"{" ,"}" ,"|" ,"~", '"', "“", "”"]
	
	for symbol in symbols:
		string = string.replace(symbol, " ")

	return string

def process_mentions(string):
	pat_mentions = re.compile("@[A-zÀ-ú0-9_]{1,15}")
	mentions = pat_mentions.findall(string)
	return [pat_mentions.sub(' ', string), mentions]

def process_hashtags(string):
	pat_hashtags = re.compile("#[A-zÀ-ú0-9_]{1,100}")
	hashtags = pat_hashtags.findall(string)
	hashtags = [h.lower() for h in hashtags]
	return [pat_hashtags.sub(' ', string), hashtags]

def toLowcase(string):
	return string.lower()

def normalize_spaces(string):
	return re.sub(" {2,}", " ", string).split()

def preprocess_tweet(tweet):
	[tweet, url] = process_urls(tweet)
	[tweet, emojis] = process_emojis(tweet)
	tweet = remove_newline(tweet)
	tweet = remove_punctuation(tweet)
	[tweet, mentions] = process_mentions(tweet)
	[tweet, hashtags] = process_hashtags(tweet)
	tweet = toLowcase(tweet)
	tweet = normalize_spaces(tweet)
	return [tweet, url, emojis, mentions, hashtags]

###################STOPWORDS###################

def remove_stopwords_numbers_short(doc):
	doc = [w for w in doc if (not w in stop_words and not w in STOP_WORDS) and (not w.isnumeric()) and (len(w) > 1)]

	return doc

def stem(tweet):
	stemmer = SnowballStemmer('spanish')
	return [stemmer.stem(i) for i in tweet]

################################################
################################################
################################################

def procesar():

	parties = ["Cs", "ERC", "JpC", "PNV", "PP", "PRC", "PSOE", "UP", "VOX"]
	wbCuentas = load_workbook(filename='Cuentas.xlsx')
	lista = wbCuentas['Hoja1']


	for party in parties:
		if not os.path.exists('./tweetsProcesados/' + party):
			os.mkdir('./tweetsProcesados/' + party)

	nCuentas = 121

	for politico in lista.iter_rows(min_row = 1, max_col = 3, max_row = nCuentas):
		if politico[1].value is not None: #Tiene tw
			
			if not os.path.exists('./tweetsProcesados/' + politico[2].value + '/' + politico[0].value): #Creacion carpeta
				os.mkdir('./tweetsProcesados/' + politico[2].value + '/' + politico[0].value)

			wbOriginal = load_workbook(filename='./root/' + politico[2].value + '/' + politico[0].value + '/tuits - ' + politico[0].value + '.xlsx')
			wbProcesado = Workbook()
			del wbProcesado["Sheet"]

			for mes in range(1,7):
				wsOriginal = wbOriginal[str(mes)]

				wbProcesado.create_sheet(str(mes))
				wsPre = wbProcesado[str(mes)]

				for tweet in wsOriginal.iter_rows(min_row=2):
					[preTweet, url, emojis, mentions, hashtags] = preprocess_tweet(tweet[0].value)
					preTweet = remove_stopwords_numbers_short(preTweet)
					stemWords = stem(preTweet)
					wsPre.append([tweet[0].value, tweet[1].value, tweet[2].value, 
						tweet[3].value, tweet[4].value, tweet[5].value, tweet[6].value, 
						" ".join(preTweet),  " ".join(stemWords), " ".join(url), 
						" ".join(emojis), " ".join(mentions), " ".join(hashtags)])
			wbProcesado.save('./tweetsProcesados/' + politico[2].value + '/' + politico[0].value + '/tuits - ' + politico[0].value + '.xlsx')

def tokenizer(tweet):
	tokens = preprocess_tweet(tweet)[0]
	stems = [w for w in tokens if (not w in stop_words and not w in STOP_WORDS) and (not w.isnumeric()) and (len(w) > 1)]
	return [stemmer.stem(i) for i in stems]
